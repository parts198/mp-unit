#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ozon_backend_v3.py

Требования (реализовано):
- supplies.json формируется по API (без ручной загрузки):
    /v3/supply-order/list + /v3/supply-order/get + /v1/supply-order/bundle + /v1/cluster/list
  фильтр по СЛОТУ (timeslot.from) в году/диапазоне; CANCELLED исключаем по умолчанию.

- Остатки на складах OZON получаем по API (без ручной загрузки):
    /v1/analytics/stocks  (батчами по <=100 sku)
  sku получаем:
    1) из supplies bundle_items (если include_items=True),
    2) добираем через /v3/product/info/list по offer_id (батчами по <=1000).

- Загрузка актов — единственное "ручное":
  * можно пачкой
  * если акт уже был загружен — не дублируем (по SHA256 содержимого).

- Аналитика:
  * заказы FBO+FBS по API (/v2/posting/fbo/list и /v3/posting/fbs/list), выбор всех/одного/нескольких магазинов.
  * уходимость по окнам наличия: поступление=slot_from_utc, конец=день, когда (принято-выкуплено) <= 0.
  * рекомендации: дефицит по кластеру после учёта OZON-остатка, ограничение — мой склад.

Запуск:
  pip install -r requirements_backend_v3.txt
  python ozon_backend_v3.py --host 0.0.0.0 --port 8000

Файлы рядом:
  stores.secrets.js   (ваш формат window.OZON_STORES = [...];)
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import random
import re
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Union

import requests
import openpyxl
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware

# -------------------- paths --------------------

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
ACTS_DIR = DATA_DIR / "acts"
CACHE_DIR = DATA_DIR / "cache"
DATA_DIR.mkdir(exist_ok=True)
ACTS_DIR.mkdir(exist_ok=True)
CACHE_DIR.mkdir(exist_ok=True)

STORES_SECRETS_JS = APP_DIR / "stores.secrets.js"
STORES_JSON = APP_DIR / "stores.json"  # fallback

SUPPLIES_PATH = DATA_DIR / "supplies.json"
ACTS_INDEX_PATH = DATA_DIR / "acts_parsed.json"
ACTS_DEDUP_PATH = DATA_DIR / "acts_dedup.json"
MY_STOCK_PATH = DATA_DIR / "my_stock.json"
OZON_STOCK_PATH = DATA_DIR / "ozon_stock.json"
ORDERS_LAST_PATH = DATA_DIR / "postings_cache.json"

DEFAULT_BASE_URL = "https://api-seller.ozon.ru"

SUPPLY_STATES_ALL: List[str] = [
    "DATA_FILLING",
    "READY_TO_SUPPLY",
    "ACCEPTED_AT_SUPPLY_WAREHOUSE",
    "IN_TRANSIT",
    "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
    "REPORTS_CONFIRMATION_AWAITING",
    "REPORT_REJECTED",
    "COMPLETED",
    "REJECTED_AT_SUPPLY_WAREHOUSE",
    "CANCELLED",
    "OVERDUE",
]


# -------------------- utils --------------------

def as_text(v: Any) -> str:
    return "" if v is None else str(v)

def load_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))

def save_json(path: Path, obj):
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

def parse_iso_utc(s: str) -> dt.datetime:
    return dt.datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(dt.timezone.utc)

def day_key(d: dt.datetime) -> str:
    return d.astimezone(dt.timezone.utc).strftime("%Y-%m-%d")

def year_window_utc(year: int) -> Tuple[dt.datetime, dt.datetime]:
    start = dt.datetime(year, 1, 1, 0, 0, 0, tzinfo=dt.timezone.utc)
    end = dt.datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=dt.timezone.utc)
    return start, end

def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

def chunked(lst: List[Any], n: int) -> Iterable[List[Any]]:
    for i in range(0, len(lst), n):
        yield lst[i:i+n]


# -------------------- stores.secrets.js parsing --------------------

def _strip_trailing_commas(s: str) -> str:
    return re.sub(r",(\s*[}\]])", r"\1", s)

def load_stores() -> List[Dict[str, str]]:
    stores = []
    if STORES_JSON.exists():
        obj = load_json(STORES_JSON, [])
        if isinstance(obj, list):
            stores = obj

    if (not stores) and STORES_SECRETS_JS.exists():
        raw = STORES_SECRETS_JS.read_text(encoding="utf-8", errors="replace")
        m = re.search(r"window\.OZON_STORES\s*=\s*(\[[\s\S]*?\])\s*;", raw, flags=re.MULTILINE)
        if not m:
            m = re.search(r"OZON_STORES\s*=\s*(\[[\s\S]*?\])\s*;", raw, flags=re.MULTILINE)
        if m:
            arr = _strip_trailing_commas(m.group(1))
            stores = json.loads(arr)

    out = []
    if isinstance(stores, list):
        for s in stores:
            if not isinstance(s, dict):
                continue
            cid = as_text(s.get("client_id")).strip()
            api = as_text(s.get("api_key")).strip()
            name = as_text(s.get("name") or cid).strip()
            if not cid:
                continue
            out.append({"name": name, "client_id": cid, "api_key": api})
    return out

def normalize_store_selection(store_sel: Any, stores: List[Dict[str, str]]) -> List[Dict[str, str]]:
    if store_sel is None:
        return stores
    if isinstance(store_sel, str) and store_sel.strip().lower() == "all":
        return stores
    if isinstance(store_sel, int):
        if 0 <= store_sel < len(stores):
            return [stores[store_sel]]
        raise HTTPException(status_code=400, detail="Неверный индекс магазина")
    if isinstance(store_sel, str):
        try:
            idx = int(store_sel)
            if 0 <= idx < len(stores):
                return [stores[idx]]
        except Exception:
            pass
        raise HTTPException(status_code=400, detail="store должен быть 'all', индексом или списком индексов")
    if isinstance(store_sel, list):
        idxs = []
        for x in store_sel:
            try:
                i = int(x)
            except Exception:
                continue
            if 0 <= i < len(stores):
                idxs.append(i)
        idxs = sorted(set(idxs))
        if not idxs:
            raise HTTPException(status_code=400, detail="Пустой список магазинов")
        return [stores[i] for i in idxs]
    raise HTTPException(status_code=400, detail="store должен быть 'all', индексом или списком")



def extract_items_from_ozon_response(resp):
    """
    Поддерживает варианты:
      {"result":{"postings":[...]}}
      {"result":{"items":[...]}}
      {"result":[...]}
      {"postings":[...]}
      {"items":[...]}
      [...]
    """
    if isinstance(resp, list):
        return [x for x in resp if isinstance(x, dict)]

    if not isinstance(resp, dict):
        return []

    result = resp.get("result")

    if isinstance(result, dict):
        items = result.get("postings") or result.get("items") or []
        if isinstance(items, list):
            return [x for x in items if isinstance(x, dict)]

    if isinstance(result, list):
        return [x for x in result if isinstance(x, dict)]

    items = resp.get("postings") or resp.get("items") or []
    if isinstance(items, list):
        return [x for x in items if isinstance(x, dict)]

    return []


# -------------------- Ozon HTTP client with retry/throttle --------------------

class OzonClient:
    def __init__(
        self,
        store: Dict[str, str],
        base_url: str = DEFAULT_BASE_URL,
        timeout_s: int = 90,
        min_interval_s: float = 0.35,
        max_retries: int = 7,
        backoff_base_s: float = 0.6,
        backoff_cap_s: float = 25.0,
        user_agent: str = "ozon-backend/3.0",
    ):
        self.store = store
        self.base_url = base_url.rstrip("/")
        self.timeout_s = timeout_s
        self.min_interval_s = float(min_interval_s)
        self.max_retries = int(max_retries)
        self.backoff_base_s = float(backoff_base_s)
        self.backoff_cap_s = float(backoff_cap_s)
        self._last_ts = 0.0

        if not store.get("api_key"):
            raise HTTPException(status_code=400, detail=f"У магазина '{store.get('name')}' пустой api_key в stores.secrets.js")

        self.session = requests.Session()
        self.session.headers.update(
            {
                "Client-Id": store["client_id"],
                "Api-Key": store["api_key"],
                "Content-Type": "application/json",
                "Accept": "application/json",
                "User-Agent": user_agent,
            }
        )

    def _throttle(self):
        if self.min_interval_s <= 0:
            return
        now = time.time()
        dt_ = now - self._last_ts
        if dt_ < self.min_interval_s:
            time.sleep(self.min_interval_s - dt_)

    def _backoff(self, attempt: int, retry_after: Optional[str] = None) -> float:
        if retry_after:
            try:
                ra = float(retry_after)
                if ra > 0:
                    return min(ra, self.backoff_cap_s)
            except Exception:
                pass
        base = self.backoff_base_s * (2 ** attempt)
        jitter = random.uniform(0, self.backoff_base_s)
        return min(base + jitter, self.backoff_cap_s)

    def post(self, path: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        url = f"{self.base_url}{path}"
        for attempt in range(self.max_retries + 1):
            self._throttle()
            r = None
            try:
                r = self.session.post(url, json=payload, timeout=self.timeout_s)
                self._last_ts = time.time()
            except requests.RequestException as e:
                if attempt >= self.max_retries:
                    raise HTTPException(status_code=502, detail=f"Ошибка сети {path}: {e}")
                time.sleep(self._backoff(attempt))
                continue

            txt = r.text or ""
            try:
                j = r.json() if txt else {}
            except Exception:
                j = {"_raw": txt[:2000]}

            if r.status_code == 429:
                if attempt >= self.max_retries:
                    raise HTTPException(status_code=429, detail=f"Лимит запросов (429) {path}: {j}")
                time.sleep(self._backoff(attempt, r.headers.get("Retry-After")))
                continue

            if r.status_code in (408, 500, 502, 503, 504):
                if attempt >= self.max_retries:
                    raise HTTPException(status_code=502, detail=f"HTTP {r.status_code} {path}: {j}")
                time.sleep(self._backoff(attempt, r.headers.get("Retry-After")))
                continue

            if not r.ok:
                raise HTTPException(status_code=502, detail=f"HTTP {r.status_code} {path}: {j}")

            return j

        raise HTTPException(status_code=502, detail=f"Не удалось выполнить {path} после ретраев")


# -------------------- Supply orders (FBO supplies) --------------------

def get_timeslot_from(order: Dict[str, Any]) -> Optional[dt.datetime]:
    ts = (order.get("timeslot") or {}).get("timeslot") or {}
    s = as_text(ts.get("from")).strip()
    if not s:
        return None
    try:
        return parse_iso_utc(s)
    except Exception:
        return None

def supply_order_list(client: OzonClient, states: List[str], limit: int, sort_by: str, sort_dir: str, last_id: Optional[str]) -> Dict[str, Any]:
    return client.post("/v3/supply-order/list", {
        "filter": {"states": states},
        "last_id": last_id,
        "limit": int(limit),
        "sort_by": sort_by,
        "sort_dir": sort_dir,
    })

def supply_order_get(client: OzonClient, order_ids: List[int]) -> Dict[str, Any]:
    return client.post("/v3/supply-order/get", {"order_ids": order_ids})

def iter_supply_order_ids(client: OzonClient, states: List[str], limit: int = 100, max_pages: Optional[int] = None) -> Iterable[int]:
    last_id: Optional[str] = None
    page = 0
    while True:
        page += 1
        if max_pages is not None and page > max_pages:
            return
        data = supply_order_list(client, states=states, limit=limit, sort_by="ORDER_CREATION", sort_dir="DESC", last_id=last_id)
        ids = data.get("order_ids") or []
        new_last = data.get("last_id")
        if not ids:
            return
        for x in ids:
            try:
                yield int(x)
            except Exception:
                continue
        if not new_last or str(new_last) == str(last_id):
            return
        last_id = str(new_last)

def fetch_supply_orders_details(client: OzonClient, order_ids: List[int], batch_size: int = 50) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for batch in chunked(order_ids, batch_size):
        resp = supply_order_get(client, batch)
        orders = resp.get("orders") or []
        if isinstance(orders, list):
            out.extend([o for o in orders if isinstance(o, dict)])
    return out

def cluster_list(client: OzonClient, cluster_type: str = "CLUSTER_TYPE_OZON") -> Dict[str, Any]:
    return client.post("/v1/cluster/list", {"cluster_type": cluster_type})

def build_wh2cluster_map(cluster_resp: Dict[str, Any]) -> Dict[int, Dict[str, Any]]:
    out: Dict[int, Dict[str, Any]] = {}
    clusters = cluster_resp.get("clusters") or []
    if not isinstance(clusters, list):
        return out
    for cl in clusters:
        if not isinstance(cl, dict):
            continue
        cl_id = cl.get("id")
        cl_name = cl.get("name")
        macro = cl.get("macrolocal_cluster_id")
        for lc in (cl.get("logistic_clusters") or []):
            if not isinstance(lc, dict):
                continue
            for wh in (lc.get("warehouses") or []):
                if not isinstance(wh, dict):
                    continue
                try:
                    wid = int(wh.get("warehouse_id"))
                except Exception:
                    continue
                out[wid] = {
                    "cluster_id": cl_id,
                    "cluster_name": cl_name,
                    "macrolocal_cluster_id": macro,
                    "warehouse_id": wid,
                    "warehouse_name": wh.get("name"),
                    "warehouse_type": wh.get("type"),
                }
    return out

def supply_order_bundle_page(client: OzonClient, bundle_id: str, dropoff_warehouse_id: int, storage_warehouse_ids: List[int], last_id: Optional[str], limit: int = 100) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "bundle_ids": [bundle_id],
        "is_asc": True,
        "limit": int(limit),
        "sort_field": "UNSPECIFIED",
        "item_tags_calculation": {
            "dropoff_warehouse_id": int(dropoff_warehouse_id),
            "storage_warehouse_ids": [int(x) for x in storage_warehouse_ids],
        },
    }
    if last_id:
        payload["last_id"] = last_id
    return client.post("/v1/supply-order/bundle", payload)

def supply_order_bundle_all(client: OzonClient, bundle_id: str, dropoff_warehouse_id: int, storage_warehouse_ids: List[int], limit: int = 100, page_delay_s: float = 0.6) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    last_id: Optional[str] = None
    while True:
        resp = supply_order_bundle_page(client, bundle_id, dropoff_warehouse_id, storage_warehouse_ids, last_id, limit=limit)
        page_items = resp.get("items") or []
        if isinstance(page_items, list):
            items.extend([x for x in page_items if isinstance(x, dict)])
        if not resp.get("has_next"):
            break
        last_id = resp.get("last_id") or None
        if not last_id:
            break
        if page_delay_s > 0:
            time.sleep(page_delay_s)
    return items

def fetch_supplies_via_api(
    store: Dict[str, str],
    year: Optional[int],
    slot_since: Optional[str],
    slot_to: Optional[str],
    include_items: bool,
    include_clusters: bool,
    exclude_cancelled: bool,
    limit: int,
    max_pages: Optional[int],
    min_interval: float,
    bundle_delay: float,
) -> Dict[str, Any]:
    client = OzonClient(store, min_interval_s=min_interval)

    if year is not None:
        s_dt, e_dt = year_window_utc(int(year))
    else:
        if not slot_since or not slot_to:
            raise HTTPException(status_code=400, detail="Нужно year либо slot_since+slot_to")
        s_dt = dt.datetime.fromisoformat(slot_since + "T00:00:00+00:00")
        e_dt = dt.datetime.fromisoformat(slot_to + "T00:00:00+00:00")

    states = [s for s in SUPPLY_STATES_ALL if (not exclude_cancelled or s != "CANCELLED")]

    order_ids = list(iter_supply_order_ids(client, states=states, limit=limit, max_pages=max_pages))
    orders = fetch_supply_orders_details(client, order_ids)

    wh2cl = {}
    if include_clusters:
        wh2cl = build_wh2cluster_map(cluster_list(client))

    filtered_orders: List[Dict[str, Any]] = []
    supplies_flat: List[Dict[str, Any]] = []

    for o in orders:
        if exclude_cancelled and as_text(o.get("state")).upper() == "CANCELLED":
            continue

        ts_from = get_timeslot_from(o)
        if ts_from is None:
            continue
        if not (s_dt <= ts_from < e_dt):
            continue

        # supplies
        supplies = o.get("supplies") or []
        if not isinstance(supplies, list):
            continue
        new_supplies = []
        for s in supplies:
            if not isinstance(s, dict):
                continue
            if exclude_cancelled and as_text(s.get("state")).upper() == "CANCELLED":
                continue
            new_supplies.append(s)

        if not new_supplies:
            continue
        o["supplies"] = new_supplies
        o["_slot_from_utc"] = ts_from.isoformat().replace("+00:00", "Z")
        o["_store"] = store.get("name")
        o["_client_id"] = store.get("client_id")

        # clusters
        if include_clusters:
            drop = o.get("drop_off_warehouse")
            if isinstance(drop, dict):
                try:
                    did = int(drop.get("warehouse_id"))
                except Exception:
                    did = None
                if did is not None:
                    drop["cluster"] = wh2cl.get(did)

            for s in new_supplies:
                storage = s.get("storage_warehouse")
                if isinstance(storage, dict):
                    try:
                        sid = int(storage.get("warehouse_id"))
                    except Exception:
                        sid = None
                    if sid is not None:
                        storage["cluster"] = wh2cl.get(sid)

        # items
        if include_items:
            drop = o.get("drop_off_warehouse") or {}
            try:
                drop_id = int(drop.get("warehouse_id"))
            except Exception:
                drop_id = None

            for s in new_supplies:
                storage = s.get("storage_warehouse") or {}
                try:
                    storage_id = int(storage.get("warehouse_id"))
                except Exception:
                    storage_id = None
                bundle_id = as_text(s.get("bundle_id")).strip()
                if not bundle_id or drop_id is None or storage_id is None:
                    continue
                items = supply_order_bundle_all(client, bundle_id, drop_id, [storage_id], limit=100, page_delay_s=bundle_delay)
                s["bundle_items"] = items

        filtered_orders.append(o)

        for s in new_supplies:
            supplies_flat.append({
                "store": store.get("name"),
                "client_id": store.get("client_id"),
                "order_id": o.get("order_id"),
                "order_number": o.get("order_number"),
                "order_state": o.get("state"),
                "order_state_updated_date": o.get("state_updated_date") or o.get("stateUpdatedDate"),
                "slot_from_utc": o.get("_slot_from_utc"),
                "drop_off_warehouse": o.get("drop_off_warehouse"),
                "supply": s,
            })

    return {
        "meta": {
            "store": store.get("name"),
            "client_id": store.get("client_id"),
            "slot_since_utc": s_dt.isoformat().replace("+00:00", "Z"),
            "slot_to_utc": e_dt.isoformat().replace("+00:00", "Z"),
            "exclude_cancelled": exclude_cancelled,
            "include_items": include_items,
            "include_clusters": include_clusters,
        },
        "orders_count": len(filtered_orders),
        "orders": filtered_orders,
        "supplies_flat_count": len(supplies_flat),
        "supplies_flat": supplies_flat,
    }


# -------------------- Product info list (offer_id -> sku) --------------------

def product_info_list_v3(client: OzonClient, offer_ids: List[str]) -> List[Dict[str, Any]]:
    resp = client.post("/v3/product/info/list", {"offer_id": offer_ids})
    return extract_items_from_ozon_response(resp)


def analytics_stocks(client: OzonClient, skus: List[int]) -> List[Dict[str, Any]]:
    resp = client.post("/v1/analytics/stocks", {"skus": [int(x) for x in skus]})
    return extract_items_from_ozon_response(resp)


def fetch_ozon_stock_via_api(
    store_sel: Any,
    offer_ids: List[str],
    offer_to_sku_from_supplies: Dict[str, int],
    min_interval: float,
) -> Dict[str, Any]:
    stores_all = load_stores()
    selected = normalize_store_selection(store_sel, stores_all)
    st = selected[0]  # stock API — достаточно по одному магазину

    client = OzonClient(st, min_interval_s=min_interval)

    # sku map: сначала из supplies, потом добираем через product/info/list
    offer_ids = sorted(set([x for x in offer_ids if x]))
    offer_to_sku = dict(offer_to_sku_from_supplies)
    missing = [o for o in offer_ids if o not in offer_to_sku]

    if missing:
        for batch in chunked(missing, 1000):
            items = product_info_list_v3(client, batch)
            for it in items:
                offer = as_text(it.get("offer_id") or it.get("offerId") or "").strip()
                sku = it.get("sku")
                try:
                    sku_i = int(sku)
                except Exception:
                    sku_i = None
                if offer and sku_i is not None:
                    offer_to_sku[offer] = sku_i

    skus = sorted(set(offer_to_sku.values()))
    if not skus:
        return {"meta": {"stores": [s["name"] for s in selected], "skus": 0}, "by_cluster": {}, "by_warehouse": {}}

    by_cluster = defaultdict(lambda: defaultdict(int))
    by_warehouse = defaultdict(lambda: defaultdict(int))
    raw_count = 0

    # reverse sku->offer
    sku_to_offer = defaultdict(list)
    for o, sku in offer_to_sku.items():
        sku_to_offer[int(sku)].append(o)

    for sku_batch in chunked(skus, 100):
        items = analytics_stocks(client, sku_batch)
        for it in items:
            raw_count += 1
            cluster_name = as_text(it.get("cluster_name") or it.get("clusterName") or "—").strip() or "—"
            warehouse_name = as_text(it.get("warehouse_name") or it.get("warehouseName") or "—").strip() or "—"
            offer = as_text(it.get("offer_id") or it.get("offerId") or "").strip()

            # иногда offer_id может не прийти, тогда восстанавливаем по sku
            if not offer:
                try:
                    sku_i = int(it.get("sku"))
                except Exception:
                    sku_i = None
                if sku_i is not None:
                    lst = sku_to_offer.get(sku_i)
                    offer = lst[0] if lst else ""

            if not offer:
                continue

            avail = it.get("available_stock_count") or it.get("availableStockCount") or 0
            try:
                avail_i = int(avail)
            except Exception:
                avail_i = 0

            by_cluster[cluster_name][offer] += avail_i
            by_warehouse[warehouse_name][offer] += avail_i

    obj = {
        "meta": {
            "fetched_at_utc": dt.datetime.now(tz=dt.timezone.utc).isoformat().replace("+00:00", "Z"),
            "store": st.get("name"),
            "client_id": st.get("client_id"),
            "offers": len(offer_ids),
            "skus": len(skus),
            "raw_items": raw_count,
        },
        "by_cluster": {k: dict(v) for k, v in by_cluster.items()},
        "by_warehouse": {k: dict(v) for k, v in by_warehouse.items()},
    }
    save_json(OZON_STOCK_PATH, obj)
    return obj


# -------------------- Postings (orders) FBO+FBS --------------------

def pick_cluster(posting: Dict[str, Any]) -> Tuple[str, str]:
    ad = posting.get("analytics_data") or posting.get("analyticsData") or {}
    fd = posting.get("financial_data") or posting.get("financialData") or {}
    cluster_from = as_text(fd.get("cluster_from") or fd.get("clusterFrom") or ad.get("cluster_from") or ad.get("clusterFrom") or posting.get("cluster_from") or posting.get("clusterFrom")).strip()
    cluster_to   = as_text(fd.get("cluster_to")   or fd.get("clusterTo")   or ad.get("cluster_to")   or ad.get("clusterTo")   or posting.get("cluster_to")   or posting.get("clusterTo")).strip()
    return cluster_from, cluster_to

def posting_date(posting: Dict[str, Any]) -> Optional[dt.datetime]:
    s = as_text(
        posting.get("created_at") or posting.get("createdAt") or
        posting.get("in_process_at") or posting.get("inProcessAt") or
        posting.get("shipment_date") or posting.get("shipmentDate") or ""
    )
    if not s:
        return None
    try:
        return parse_iso_utc(s)
    except Exception:
        return None

def posting_status_bucket(status: str) -> str:
    s = (status or "").strip().lower()
    if "cancel" in s or "отмен" in s:
        return "Отменено"
    if "deliver" in s or "достав" in s:
        return "Доставлено"
    return "В процессе"

def iter_posting_lines(posting: Dict[str, Any]) -> List[Dict[str, Any]]:
    items = posting.get("products") or posting.get("items") or posting.get("posting_items") or posting.get("posting_items_v2") or []
    if not isinstance(items, list):
        return []
    return [x for x in items if isinstance(x, dict)]

def ozon_postings_list(client: OzonClient, schema: str, since_iso: str, to_iso: str, limit: int = 1000) -> List[Dict[str, Any]]:
    path = "/v3/posting/fbs/list" if schema == "FBS" else "/v2/posting/fbo/list"
    out = []
    offset = 0
    for _ in range(0, 400):
        body = {
            "dir": "ASC",
            "filter": {"since": since_iso, "to": to_iso},
            "limit": limit,
            "offset": offset,
            "with": {"analytics_data": True, "financial_data": True},
        }
        resp = client.post(path, body)
        items = extract_items_from_ozon_response(resp)
        if not items:
            break
        for it in items:
            it["_schema"] = schema
            it["_store"] = client.store.get("name")
            out.append(it)
        if len(items) < limit:
            break
        offset += limit
    return out


def fetch_postings_cached(date_from: str, date_to: str, schemas: List[str], store_sel: Any, min_interval: float) -> Dict[str, Any]:
    stores_all = load_stores()
    selected = normalize_store_selection(store_sel, stores_all)
    idxs = []
    for st in selected:
        for i, s in enumerate(stores_all):
            if s["client_id"] == st["client_id"]:
                idxs.append(i)
    idxs = sorted(set(idxs))

    schemas = [s for s in schemas if s in ("FBO", "FBS")]
    if not schemas:
        schemas = ["FBO", "FBS"]

    key = f"{date_from}_{date_to}_{','.join(sorted(schemas))}_stores-{','.join(map(str,idxs))}"
    key = re.sub(r"[^A-Za-z0-9_,\-.]", "_", key)
    cache_file = CACHE_DIR / f"postings_{key}.json"
    if cache_file.exists():
        obj = load_json(cache_file, {})
        save_json(ORDERS_LAST_PATH, obj)
        return obj

    since_iso = date_from + "T00:00:00.000Z"
    to_iso = date_to + "T23:59:59.999Z"

    postings = []
    errors = []
    for st in selected:
        c = OzonClient(st, min_interval_s=min_interval)
        for sch in schemas:
            try:
                postings.extend(ozon_postings_list(c, sch, since_iso, to_iso))
            except HTTPException as e:
                errors.append({"store": st.get("name"), "schema": sch, "detail": e.detail})

    obj = {"meta": {"date_from": date_from, "date_to": date_to, "schemas": schemas, "stores": idxs, "errors": errors}, "postings": postings}
    save_json(cache_file, obj)
    save_json(ORDERS_LAST_PATH, obj)
    return obj


# -------------------- Acts parsing + dedup --------------------

def find_supply_id_from_filename(name: str) -> Optional[str]:
    m = re.search(r"supply-(\d+)-acceptance-report", name, flags=re.IGNORECASE)
    return m.group(1) if m else None

def read_acceptance_report_xlsx_bytes(content: bytes) -> Dict[str, Dict[str, Any]]:
    # читаем из временного файла (openpyxl не умеет bytes напрямую надежно)
    tmp = DATA_DIR / f"_tmp_act_{random.randint(1, 10**9)}.xlsx"
    tmp.write_bytes(content)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
        ws = wb["Отчёт по поставке"] if "Отчёт по поставке" in wb.sheetnames else wb[wb.sheetnames[0]]

        header_row = None
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and len(row) >= 6 and row[0] == "Ozon ID" and row[2] == "Артикул товара":
                header_row = i
                headers = list(row)
                break
        if header_row is None or headers is None:
            raise RuntimeError("Не найден заголовок таблицы (Ozon ID / Артикул товара ...)")

        def idx(col: str) -> int:
            try:
                return headers.index(col)
            except ValueError:
                return -1

        i_art = idx("Артикул товара")
        i_decl = idx("Заявлено (шт.)")
        i_acc = idx("Принято на склад (шт.)")

        if min(i_art, i_decl, i_acc) < 0:
            raise RuntimeError("В акте не хватает ожидаемых колонок.")

        out: Dict[str, Dict[str, Any]] = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(v is None for v in row):
                if out:
                    break
                continue
            offer = as_text(row[i_art]).strip()
            if not offer:
                continue
            declared = row[i_decl] if i_decl < len(row) else 0
            accepted = row[i_acc] if i_acc < len(row) else 0
            try: declared_i = int(declared) if declared is not None else 0
            except Exception: declared_i = 0
            try: accepted_i = int(accepted) if accepted is not None else 0
            except Exception: accepted_i = 0
            out.setdefault(offer, {"declared": 0, "accepted": 0})
            out[offer]["declared"] += declared_i
            out[offer]["accepted"] += accepted_i
        return out
    finally:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass

def rebuild_acts_index() -> Dict[str, Any]:
    idx: Dict[str, Any] = {"by_supply": {}, "files": []}
    for f in sorted(ACTS_DIR.glob("*.xlsx")):
        sid = find_supply_id_from_filename(f.name)
        if not sid:
            continue
        try:
            items = read_acceptance_report_xlsx_bytes(f.read_bytes())
        except Exception as e:
            items = {"_error": str(e)}
        idx["by_supply"][sid] = items
        idx["files"].append(f.name)
    save_json(ACTS_INDEX_PATH, idx)
    return idx

def load_my_stock() -> Dict[str, int]:
    return load_json(MY_STOCK_PATH, {})

def save_my_stock_from_upload(path: Path) -> Dict[str, int]:
    # простой CSV/XLSX парсер: offer_id + qty
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = None
        for row in rows:
            if row and any(v is not None and str(v).strip() for v in row):
                headers = [as_text(v).strip() for v in row]
                break
        if not headers:
            return {}
        def find_col(cands):
            for i, h in enumerate(headers):
                h0 = h.lower()
                if any(k in h0 for k in cands):
                    return i
            return None
        i_offer = find_col(["offer", "артикул"])
        i_qty = find_col(["qty", "остат", "колич", "quantity"])
        if i_offer is None or i_qty is None:
            raise RuntimeError(f"Не нашёл колонки offer/артикул и qty/остаток. Заголовки: {headers}")
        out = defaultdict(int)
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            offer = as_text(row[i_offer]).strip() if i_offer < len(row) else ""
            if not offer:
                continue
            q = row[i_qty] if i_qty < len(row) else 0
            try: qty = int(float(str(q).replace(",", ".")))
            except Exception: qty = 0
            out[offer] += qty
        stock = dict(out)
        save_json(MY_STOCK_PATH, stock)
        return stock

    # CSV
    txt = path.read_text(encoding="utf-8-sig", errors="replace")
    lines = [ln for ln in txt.splitlines() if ln.strip()]
    if not lines:
        return {}
    delim = ";" if lines[0].count(";") >= lines[0].count(",") else ","
    import csv
    reader = csv.DictReader(lines, delimiter=delim)
    cols = reader.fieldnames or []
    def find_col(cands):
        for c in cols:
            c0 = c.strip().lower()
            if any(k in c0 for k in cands):
                return c
        return None
    col_offer = find_col(["offer", "артикул"])
    col_qty = find_col(["qty", "остат", "колич", "quantity"])
    if not col_offer or not col_qty:
        raise RuntimeError(f"Не нашёл колонки offer/артикул и qty/остаток. Колонки: {cols}")

    out = defaultdict(int)
    for r in reader:
        offer = as_text(r.get(col_offer)).strip()
        if not offer:
            continue
        q = r.get(col_qty)
        try: qty = int(float(str(q).replace(",", ".")))
        except Exception: qty = 0
        out[offer] += qty
    stock = dict(out)
    save_json(MY_STOCK_PATH, stock)
    return stock


# -------------------- Analytics (velocity + recommendations) --------------------

def build_inflows_from_supplies_and_acts(supplies: Dict[str, Any], acts_idx: Dict[str, Any]) -> List[Dict[str, Any]]:
    by_supply = (acts_idx or {}).get("by_supply") or {}
    supplies_flat = supplies.get("supplies_flat") or []
    out = []
    for rec in supplies_flat:
        supply = rec.get("supply") or {}
        sid = as_text(supply.get("supply_id")).strip()
        if not sid:
            continue
        slot_from = as_text(rec.get("slot_from_utc")).strip()
        if not slot_from:
            continue
        try:
            d = day_key(parse_iso_utc(slot_from))
        except Exception:
            continue

        storage = supply.get("storage_warehouse") or {}
        cl_obj = storage.get("cluster") or {}
        cl_name = as_text(cl_obj.get("cluster_name")).strip() if isinstance(cl_obj, dict) else ""
        cl = cl_name or "—"

        act_items = by_supply.get(sid) or {}
        if isinstance(act_items, dict) and "_error" in act_items:
            act_items = {}
        if not isinstance(act_items, dict):
            act_items = {}

        for offer_id, v in act_items.items():
            if not isinstance(v, dict):
                continue
            qty = int(v.get("accepted", 0) or 0)
            if qty <= 0:
                continue
            out.append({
                "date": d,
                "cluster_storage": cl,
                "offer_id": str(offer_id),
                "qty": qty,
                "supply_id": sid,
            })
    return out

def build_outflows_from_postings(postings_obj: Dict[str, Any], mode: str) -> List[Dict[str, Any]]:
    postings = postings_obj.get("postings") or []
    out = []
    for p in postings:
        if not isinstance(p, dict):
            continue
        dtp = posting_date(p)
        if dtp is None:
            continue
        d = day_key(dtp)
        cl_from, cl_to = pick_cluster(p)
        cl = (cl_from if mode == "from" else cl_to).strip() or "—"

        status = as_text(p.get("status")).strip()
        bucket = posting_status_bucket(status)

        for line in iter_posting_lines(p):
            offer = as_text(line.get("offer_id") or line.get("offerId") or "").strip()
            if not offer:
                continue
            q = line.get("quantity") or line.get("qty") or line.get("count") or 0
            try:
                qty = int(q)
            except Exception:
                qty = 0
            if qty <= 0:
                continue
            out.append({"date": d, "cluster": cl, "offer_id": offer, "qty": qty, "bucket": bucket})
    return out

def build_demand_stats_from_postings(postings_obj: Dict[str, Any], mode: str = 'to') -> Dict[Tuple[str, str], Dict[str, Any]]:
    events = build_outflows_from_postings(postings_obj, mode=mode)

    by_key = defaultdict(lambda: {
        'cluster': '',
        'offer_id': '',
        'total_delivered': 0,
        'first_delivery_date': None,
        'last_delivery_date': None,
    })

    for ev in events:
        if ev.get('bucket') != 'Доставлено':
            continue

        cluster = ev.get('cluster') or '—'
        offer_id = ev.get('offer_id')
        qty = int(ev.get('qty') or 0)
        d = ev.get('date')

        key = (cluster, offer_id)
        row = by_key[key]
        row['cluster'] = cluster
        row['offer_id'] = offer_id
        row['total_delivered'] += qty

        if d:
            if row['first_delivery_date'] is None or d < row['first_delivery_date']:
                row['first_delivery_date'] = d
            if row['last_delivery_date'] is None or d > row['last_delivery_date']:
                row['last_delivery_date'] = d

    out = {}
    for key, row in by_key.items():
        fd = row['first_delivery_date']
        ld = row['last_delivery_date']
        if fd and ld:
            days = (dt.date.fromisoformat(ld) - dt.date.fromisoformat(fd)).days + 1
        else:
            days = 0
        days = max(days, 1)
        row['sales_days'] = days
        row['avg_daily_delivered_sales_window'] = (row['total_delivered'] / days) if days else 0.0
        out[key] = row

    return out

def build_availability_map_from_velocity(
    vel: Dict[Tuple[str, str], Dict[str, Any]],
    date_to: str,
    ozon_stock: Dict[Tuple[str, str], int]
) -> Dict[Tuple[str, str], Dict[str, Any]]:
    out = {}
    for key, v in vel.items():
        start_s = v.get('first_inflow_date')
        if not start_s:
            continue

        start_d = dt.date.fromisoformat(start_s)
        cur_ozon = int(ozon_stock.get(key, 0) or 0)

        if cur_ozon > 0:
            end_s = date_to
        else:
            end_s = v.get('last_depletion_date') or date_to

        end_d = dt.date.fromisoformat(end_s)
        days = (end_d - start_d).days + 1
        days = max(days, 1)

        out[key] = {
            'start_date': start_s,
            'end_date': end_s,
            'days_active': days,
            'current_ozon': cur_ozon,
        }

    return out

def compute_velocity_windows(inflows: List[Dict[str, Any]], outflows: List[Dict[str, Any]], date_from: str, date_to: str) -> Dict[Tuple[str, str], Dict[str, Any]]:
    inflow_map = defaultdict(int)  # (cluster, offer, date) -> qty
    for ev in inflows:
        inflow_map[(ev["cluster_storage"], ev["offer_id"], ev["date"])] += int(ev["qty"])

    out_map = defaultdict(int)  # (cluster, offer, date) -> qty (delivered only)
    for ev in outflows:
        if ev.get("bucket") != "Доставлено":
            continue
        out_map[(ev["cluster"], ev["offer_id"], ev["date"])] += int(ev["qty"])

    keys = set((k[0], k[1]) for k in inflow_map.keys()) | set((k[0], k[1]) for k in out_map.keys())

    start = dt.date.fromisoformat(date_from)
    end = dt.date.fromisoformat(date_to)

    def daterange(a: dt.date, b: dt.date):
        cur = a
        while cur <= b:
            yield cur
            cur += dt.timedelta(days=1)

    res: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for cl, offer in keys:
        stock = 0
        in_window = False
        win_start = None
        win_sold = 0
        windows = []
        first_inflow = None
        last_depl = None

        for day in daterange(start, end):
            d = day.isoformat()
            infl = inflow_map.get((cl, offer, d), 0)
            outq = out_map.get((cl, offer, d), 0)

            if infl > 0 and first_inflow is None:
                first_inflow = d

            if (stock <= 0) and infl > 0:
                in_window = True
                win_start = d
                win_sold = 0

            stock += infl
            if stock > 0:
                sell = min(stock, outq)
                stock -= sell
                win_sold += sell

            if in_window and stock <= 0:
                in_window = False
                win_end = d
                days = (dt.date.fromisoformat(win_end) - dt.date.fromisoformat(win_start)).days + 1 if win_start else 1
                windows.append({"start": win_start, "end": win_end, "days": days, "sold": win_sold})
                last_depl = win_end
                win_start = None
                win_sold = 0

        if in_window and win_start:
            days = (end - dt.date.fromisoformat(win_start)).days + 1
            windows.append({"start": win_start, "end": date_to, "days": days, "sold": win_sold})

        total_days = sum(w["days"] for w in windows) or 0
        total_sold = sum(w["sold"] for w in windows) or 0
        avg = (total_sold / total_days) if total_days else 0.0

        res[(cl, offer)] = {
            "cluster": cl,
            "offer_id": offer,
            "windows": windows,
            "total_days": total_days,
            "total_sold": total_sold,
            "avg_daily_sold": avg,
            "last_stock_est": stock if stock > 0 else 0,
            "first_inflow_date": first_inflow,
            "last_depletion_date": last_depl,
        }
    return res

def ozon_cluster_stock_map() -> Dict[Tuple[str, str], int]:
    oz = load_json(OZON_STOCK_PATH, {"by_cluster": {}})
    by_cluster = oz.get("by_cluster") or {}
    out = {}
    if isinstance(by_cluster, dict):
        for cl, m in by_cluster.items():
            if not isinstance(m, dict):
                continue
            for offer, qty in m.items():
                try:
                    out[(str(cl), str(offer))] = int(qty)
                except Exception:
                    pass
    return out

def build_recommendations(
    demand_stats: Dict[Tuple[str, str], Dict[str, Any]],
    availability_map: Dict[Tuple[str, str], Dict[str, Any]],
    date_to: str,
    cover_days: int,
    my_stock: Dict[str, int],
    ozon_stock: Dict[Tuple[str, str], int]
) -> List[Dict[str, Any]]:
    by_offer = defaultdict(list)

    for (cl, offer), d in demand_stats.items():
        total_delivered = int(d.get('total_delivered', 0) or 0)
        if total_delivered <= 0:
            continue

        cur_ozon = int(ozon_stock.get((cl, offer), 0) or 0)
        avail = availability_map.get((cl, offer))

        if avail:
            start_s = avail['start_date']
            end_s = avail['end_date']
            days_active = int(avail['days_active'])
        else:
            start_s = d.get('first_delivery_date')
            if not start_s:
                continue
            if cur_ozon > 0:
                end_s = date_to
            else:
                end_s = d.get('last_delivery_date') or date_to
            start_d = dt.date.fromisoformat(start_s)
            end_d = dt.date.fromisoformat(end_s)
            days_active = max((end_d - start_d).days + 1, 1)

        avg = float(total_delivered) / float(days_active)
        target = avg * float(cover_days)
        rec = max(0.0, target - cur_ozon)

        if rec <= 0:
            continue

        by_offer[offer].append((cl, d, rec, cur_ozon, avg, start_s, end_s, days_active, total_delivered, target))

    rows = []
    for offer, lst in by_offer.items():
        total_rec = sum(x[2] for x in lst)
        if total_rec <= 0:
            continue

        limit = int(my_stock.get(offer, 0) or 0)

        for cl, d, rec, cur_ozon, avg, start_s, end_s, days_active, total_delivered, target in lst:
            capped = 0
            reason = ''
            if limit <= 0:
                reason = 'Нет остатка на вашем складе (не загружен или 0)'
            else:
                share = (rec / total_rec) if total_rec else 0.0
                capped = int(round(limit * share))
                reason = f'Остаток на вашем складе: {limit}'

            rows.append({
                'cluster': cl,
                'offer_id': offer,
                'avg_daily': round(avg, 4),
                'cover_days': int(cover_days),
                'target': round(target, 2),
                'current_ozon': int(cur_ozon),
                'current_est': int(cur_ozon),
                'recommended': int(round(rec)),
                'recommended_capped': int(capped),
                'cap_reason': reason,
                'first_inflow_date': start_s,
                'last_depletion_date': end_s,
                'window_start': start_s,
                'window_end': end_s,
                'total_days': int(days_active),
                'total_sold': int(total_delivered),
                'first_delivery_date': d.get('first_delivery_date'),
                'last_delivery_date': d.get('last_delivery_date'),
            })

    rows.sort(
        key=lambda x: (
            x.get('recommended_capped', 0),
            x.get('recommended', 0),
            x.get('avg_daily', 0),
            x.get('total_sold', 0),
        ),
        reverse=True
    )
    return rows



# -------------------- FastAPI --------------------

app = FastAPI(title="Ozon Analytics Backend", version="3.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/api/state")
def api_state():
    stores = load_stores()
    acts_idx = load_json(ACTS_INDEX_PATH, {"files": []})
    dedup = load_json(ACTS_DEDUP_PATH, {"hashes": {}})
    my_stock = load_my_stock()
    ozon_stock = load_json(OZON_STOCK_PATH, {"by_cluster": {}})
    orders_meta = load_json(ORDERS_LAST_PATH, {}).get("meta")
    supplies_meta = load_json(SUPPLIES_PATH, {}).get("meta")
    return {
        "stores": [{"name": s["name"]} for s in stores],
        "supplies_meta": supplies_meta,
        "acts_count": len((acts_idx or {}).get("files") or []),
        "acts_dedup_hashes": len((dedup or {}).get("hashes") or {}),
        "acts_dedup_supply_ids": len((dedup or {}).get("supply_ids") or {}),
        "my_stock_items": len(my_stock or {}),
        "ozon_stock_clusters": len((ozon_stock or {}).get("by_cluster") or {}),
        "last_orders_fetch": orders_meta,
        "ozon_stock_meta": (ozon_stock or {}).get("meta"),
    }

@app.post("/api/upload/acts")
async def api_upload_acts(files: List[UploadFile] = File(...)):
    dedup = load_json(ACTS_DEDUP_PATH, {"hashes": {}, "supply_ids": {}})
    hashes: Dict[str, str] = dedup.get("hashes") or {}
    supply_ids: Dict[str, str] = dedup.get("supply_ids") or {}
    saved, skipped = [], []

    for f in files:
        if not f.filename.lower().endswith(".xlsx"):
            skipped.append({"file": f.filename, "reason": "не .xlsx"})
            continue
        content = await f.read()
        h = sha256_bytes(content)
        if h in hashes:
            skipped.append({"file": f.filename, "reason": "уже загружен (sha256)", "as": hashes[h]})
            continue

        # сохраняем
        out = ACTS_DIR / Path(f.filename).name
        # если файл с таким именем уже есть, добавим суффикс
        if out.exists():
            out = ACTS_DIR / (out.stem + f"__{h[:8]}" + out.suffix)
        out.write_bytes(content)
        hashes[h] = out.name
        if sid:
            supply_ids[sid] = out.name
        saved.append(out.name)

    save_json(ACTS_DEDUP_PATH, {"hashes": hashes, "supply_ids": supply_ids})
    idx = rebuild_acts_index()
    return {"ok": True, "saved": saved, "skipped": skipped, "acts_count": len(idx.get("files") or [])}

@app.post("/api/upload/my_stock")
async def api_upload_my_stock(file: UploadFile = File(...)):
    name = file.filename.lower()
    content = await file.read()
    tmp = DATA_DIR / ("_upload_" + Path(file.filename).name)
    tmp.write_bytes(content)
    try:
        stock = save_my_stock_from_upload(tmp)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Не смогли прочитать остатки: {e}")
    return {"ok": True, "items": len(stock)}

@app.post("/api/fetch/supplies")
def api_fetch_supplies(payload: Dict[str, Any]):
    year = payload.get("year")
    slot_since = payload.get("slot_since")
    slot_to = payload.get("slot_to")
    store_sel = payload.get("store", "all")
    include_items = bool(payload.get("include_items", True))
    include_clusters = bool(payload.get("include_clusters", True))
    exclude_cancelled = bool(payload.get("exclude_cancelled", True))
    limit = int(payload.get("limit", 100))
    max_pages = payload.get("max_pages")
    min_interval = float(payload.get("min_interval", 0.35))
    bundle_delay = float(payload.get("bundle_delay", 0.6))

    stores_all = load_stores()
    selected = normalize_store_selection(store_sel, stores_all)

    merged_orders = []
    merged_flat = []
    metas = []
    for st in selected:
        obj = fetch_supplies_via_api(
            store=st,
            year=int(year) if year is not None else None,
            slot_since=slot_since,
            slot_to=slot_to,
            include_items=include_items,
            include_clusters=include_clusters,
            exclude_cancelled=exclude_cancelled,
            limit=limit,
            max_pages=int(max_pages) if max_pages is not None else None,
            min_interval=min_interval,
            bundle_delay=bundle_delay,
        )
        metas.append(obj.get("meta"))
        merged_orders.extend(obj.get("orders") or [])
        merged_flat.extend(obj.get("supplies_flat") or [])

    supplies = {
        "meta": {
            "stores": [s["name"] for s in selected],
            "generated_at_utc": dt.datetime.now(tz=dt.timezone.utc).isoformat().replace("+00:00", "Z"),
            "parts": metas,
        },
        "orders_count": len(merged_orders),
        "orders": merged_orders,
        "supplies_flat_count": len(merged_flat),
        "supplies_flat": merged_flat,
    }
    save_json(SUPPLIES_PATH, supplies)

    return {"ok": True, "orders_count": len(merged_orders), "supplies_flat_count": len(merged_flat)}

@app.post("/api/fetch/postings")
def api_fetch_postings(payload: Dict[str, Any]):
    date_from = as_text(payload.get("date_from")).strip()
    date_to = as_text(payload.get("date_to")).strip()
    schemas = payload.get("schemas") or ["FBO", "FBS"]
    store_sel = payload.get("store", "all")
    min_interval = float(payload.get("min_interval", 0.35))
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="Нужны date_from и date_to (YYYY-MM-DD)")
    obj = fetch_postings_cached(date_from, date_to, schemas, store_sel, min_interval=min_interval)
    return {"ok": True, "meta": obj.get("meta"), "postings_count": len(obj.get("postings") or [])}

@app.post("/api/fetch/ozon_stock")
def api_fetch_ozon_stock(payload: Dict[str, Any]):
    if not SUPPLIES_PATH.exists():
        raise HTTPException(status_code=400, detail="Сначала сформируйте supplies.json через /api/fetch/supplies")
    supplies = load_json(SUPPLIES_PATH, {})
    postings_obj = load_json(ORDERS_LAST_PATH, {})  # опционально
    my_stock = load_my_stock()

    store_sel = payload.get("store", "all")
    min_interval = float(payload.get("min_interval", 0.35))

    # 1) offer->sku из supplies (если есть bundle_items)
    offer_to_sku_sup: Dict[str, int] = {}
    offer_ids = set()

    for rec in (supplies.get("supplies_flat") or []):
        s = rec.get("supply") or {}
        for it in (s.get("bundle_items") or []):
            if not isinstance(it, dict):
                continue
            offer = as_text(it.get("offer_id")).strip()
            sku = it.get("sku")
            if offer:
                offer_ids.add(offer)
            try:
                sku_i = int(sku)
            except Exception:
                sku_i = None
            if offer and sku_i is not None:
                offer_to_sku_sup[offer] = sku_i

    # 2) offer_ids из актов (если товаров нет в bundle_items)
    acts_idx = load_json(ACTS_INDEX_PATH, {"by_supply": {}})
    for sid, m in (acts_idx.get("by_supply") or {}).items():
        if isinstance(m, dict) and "_error" in m:
            continue
        if isinstance(m, dict):
            for offer in m.keys():
                offer_ids.add(str(offer))

    # 3) offer_ids из заказов (если уже тянули)
    for p in (postings_obj.get("postings") or []):
        if not isinstance(p, dict):
            continue
        for line in iter_posting_lines(p):
            offer = as_text(line.get("offer_id") or line.get("offerId") or "").strip()
            if offer:
                offer_ids.add(offer)

    # 4) offer_ids из моего склада
    for offer in (my_stock or {}).keys():
        offer_ids.add(str(offer))

    offer_ids_list = sorted(offer_ids)
    obj = fetch_ozon_stock_via_api(
        store_sel=store_sel,
        offer_ids=offer_ids_list,
        offer_to_sku_from_supplies=offer_to_sku_sup,
        min_interval=min_interval,
    )
    return {
        "ok": True,
        "meta": obj.get("meta"),
        "clusters": len(obj.get("by_cluster") or {}),
        "warehouses": len(obj.get("by_warehouse") or {}),
    }

@app.post("/api/analyze")
def api_analyze(payload: Dict[str, Any]):
    if not SUPPLIES_PATH.exists():
        raise HTTPException(status_code=400, detail='Сначала сформируйте supplies.json через /api/fetch/supplies')
    supplies = load_json(SUPPLIES_PATH, {})
    acts_idx = load_json(ACTS_INDEX_PATH, {'by_supply': {}})
    postings_obj = load_json(ORDERS_LAST_PATH, {})
    if not postings_obj:
        raise HTTPException(status_code=400, detail='Сначала подтяните заказы через /api/fetch/postings')

    date_from = as_text(payload.get('date_from')).strip()
    date_to = as_text(payload.get('date_to')).strip()
    mode = as_text(payload.get('mode') or 'to').strip()
    cover_days = int(payload.get('cover_days') or 14)
    if mode not in ('from', 'to'):
        raise HTTPException(status_code=400, detail="mode должен быть 'from' или 'to'")

    inflows = build_inflows_from_supplies_and_acts(supplies, acts_idx)
    outflows_from = build_outflows_from_postings(postings_obj, mode='from')
    vel = compute_velocity_windows(inflows, outflows_from, date_from, date_to)

    demand_stats_selected = build_demand_stats_from_postings(postings_obj, mode=mode)
    demand_top = sorted(
        [
            {
                'cluster': d['cluster'],
                'offer_id': d['offer_id'],
                'delivered_qty': d['total_delivered'],
            }
            for d in demand_stats_selected.values()
        ],
        key=lambda x: x['delivered_qty'],
        reverse=True
    )[:2000]

    demand_stats_to = build_demand_stats_from_postings(postings_obj, mode='to')

    my_stock = load_my_stock()
    oz_stock = ozon_cluster_stock_map()
    availability_map = build_availability_map_from_velocity(vel, date_to=date_to, ozon_stock=oz_stock)

    recs = build_recommendations(
        demand_stats=demand_stats_to,
        availability_map=availability_map,
        date_to=date_to,
        cover_days=cover_days,
        my_stock=my_stock,
        ozon_stock=oz_stock,
    )

    vel_rows = []
    for (_, _), v in vel.items():
        if not v.get('first_inflow_date'):
            continue
        vel_rows.append({
            'cluster': v['cluster'],
            'offer_id': v['offer_id'],
            'first_inflow_date': v.get('first_inflow_date'),
            'last_depletion_date': v.get('last_depletion_date'),
            'avg_daily': round(float(v.get('avg_daily_sold', 0.0)), 4),
            'total_days': v.get('total_days'),
            'total_sold': v.get('total_sold'),
            'current_ozon': oz_stock.get((v['cluster'], v['offer_id'])),
        })
    vel_rows.sort(key=lambda x: (x['avg_daily'], x['total_sold']), reverse=True)
    vel_rows = vel_rows[:5000]

    return {
        'ok': True,
        'meta': {
            'date_from': date_from,
            'date_to': date_to,
            'mode': mode,
            'cover_days': cover_days,
            'recommendations_logic': 'cluster_to demand with availability window (arrival -> depletion/today), minus ozon stock, capped by my_stock',
            'orders_errors': (postings_obj.get('meta') or {}).get('errors') or []
        },
        'postings_count': len(postings_obj.get('postings') or []),
        'velocity_count': len(vel),
        'velocity_rows': vel_rows,
        'recommendations_count': len(recs),
        'recommendations': recs[:2000],
        'demand_top': demand_top,
    }



def main():
    import uvicorn
    ap = argparse.ArgumentParser()
    ap.add_argument("--host", default="127.0.0.1")
    ap.add_argument("--port", type=int, default=8000)
    args = ap.parse_args()
    uvicorn.run("ozon_backend_v5:app", host=args.host, port=args.port, reload=False)

if __name__ == "__main__":
    main()
